from __future__ import annotations

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods

from tracking.models import RouteCatalog, TrackPoint
from .services import (
    load_points,
    gps_filter_jumps,
    calc_total_km,
    count_sand_base_entries,
    split_trips_from_sand_base,
    slim_points,
    SAND_BASE_LAT,
    SAND_BASE_LON,
    SAND_BASE_RADIUS_KM,
)


def fmt_dt(v):
    return v  # строки dt_from/dt_to возвращаем как пришло (как в FastAPI)


@require_http_methods(["GET"])
def api_routes(request):
    qs = RouteCatalog.objects.order_by("name").values(
        "name", "road_width_m", "pss_tonnage_t", "road_length_km"
    )
    return JsonResponse({"routes": list(qs)})


@require_http_methods(["GET"])
def api_oids(request):
    oids = list(
        TrackPoint.objects.values_list("oid", flat=True).distinct().order_by("oid")
    )
    return JsonResponse({"oids": oids})


@require_http_methods(["GET"])
def points_summary(request):
    oid = int(request.GET.get("oid"))
    dt_from = request.GET.get("dt_from")
    dt_to = request.GET.get("dt_to")
    limit = int(request.GET.get("limit", "500000"))
    max_jump_km = float(request.GET.get("max_jump_km", "1.0"))
    max_speed_kmh = float(request.GET.get("max_speed_kmh", "180.0"))

    pts = load_points(oid=oid, dt_from=dt_from, dt_to=dt_to, limit=limit)
    pts_f, gps_stats = gps_filter_jumps(
        pts, max_jump_km=max_jump_km, max_speed_kmh=max_speed_kmh
    )

    total_km = calc_total_km(pts_f)
    sand_base_entries = count_sand_base_entries(pts_f)

    trips, _entry_idx = split_trips_from_sand_base(pts_f)
    trips_km_filtered = [tr for tr in trips if calc_total_km(tr) >= 1.0]

    return JsonResponse(
        {
            "oid": oid,
            "dt_from": fmt_dt(dt_from),
            "dt_to": fmt_dt(dt_to),
            "points_count_original": len(pts),
            "points_count_used": len(pts_f),
            "gps_jumps_removed": gps_stats["removed"],
            "max_jump_km": max_jump_km,
            "max_speed_kmh": max_speed_kmh,
            "total_km": total_km,
            "sand_base_entries": sand_base_entries,
            "trips_count_raw": len(trips),
            "trips_count_ge1km": len(trips_km_filtered),
        }
    )


@require_http_methods(["GET"])
def trips_for_map(request):
    oid = int(request.GET.get("oid"))
    dt_from = request.GET.get("dt_from")
    dt_to = request.GET.get("dt_to")
    limit = int(request.GET.get("limit", "500000"))
    max_points_per_trip = int(request.GET.get("max_points_per_trip", "2000"))
    max_jump_km = float(request.GET.get("max_jump_km", "1.0"))
    max_speed_kmh = float(request.GET.get("max_speed_kmh", "180.0"))
    min_trip_km = float(request.GET.get("min_trip_km", "1.0"))

    pts = load_points(oid=oid, dt_from=dt_from, dt_to=dt_to, limit=limit)
    pts_f, gps_stats = gps_filter_jumps(
        pts, max_jump_km=max_jump_km, max_speed_kmh=max_speed_kmh
    )

    trips, _ = split_trips_from_sand_base(pts_f)

    trips_out = []
    trip_no = 0
    for tr in trips:
        dist_km = calc_total_km(tr)
        if dist_km < min_trip_km:
            continue

        trip_no += 1
        slim, step = slim_points(tr, max_points=max_points_per_trip)

        points_out = [{"lat": p.lat, "lon": p.lon, "tm": p.tm} for p in slim]
        trips_out.append(
            {
                "trip_no": trip_no,
                "tm_start": tr[0].tm,
                "tm_end": tr[-1].tm,
                "original_points": len(tr),
                "step": step,
                "distance_km": dist_km,
                "points": points_out,
            }
        )

    return JsonResponse(
        {
            "oid": oid,
            "dt_from": fmt_dt(dt_from),
            "dt_to": fmt_dt(dt_to),
            "original_count": len(pts),
            "filtered_count": len(pts_f),
            "gps_jumps_removed": gps_stats["removed"],
            "sand_base_entries": count_sand_base_entries(pts_f),
            "trips_count": len(trips_out),
            "min_trip_km": min_trip_km,
            "max_points_per_trip": max_points_per_trip,
            "sand_base": {
                "lat": SAND_BASE_LAT,
                "lon": SAND_BASE_LON,
                "radius_km": SAND_BASE_RADIUS_KM,
            },
            "trips": trips_out,
        }
    )


# --- FORMS API (putevoy_forms) ---

import json
from tempfile import NamedTemporaryFile

from django.http import JsonResponse, HttpResponseBadRequest, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime

from openpyxl import load_workbook

from formsapp.models import PutevoyForm


def _parse_dt_from_input(v: str):
    # из html datetime-local приходит "YYYY-MM-DDTHH:MM"
    if not v:
        return None
    s = str(v).strip().replace("T", " ")
    if len(s) == 16:
        s += ":00"
    return parse_datetime(s)


@csrf_exempt
def forms_save(request):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    meta = payload.get("meta") or {}
    oid = meta.get("oid") or None

    form = PutevoyForm.objects.create(
        oid=int(oid) if str(oid).isdigit() else None,
        dt_from=_parse_dt_from_input(meta.get("dt_from", "")),
        dt_to=_parse_dt_from_input(meta.get("dt_to", "")),
        payload=payload,
    )
    return JsonResponse({"form_id": form.id})


def forms_list(request):
    qs = PutevoyForm.objects.all()

    oid = request.GET.get("oid")
    if oid and oid.isdigit():
        qs = qs.filter(oid=int(oid))

    limit = request.GET.get("limit", "50")
    try:
        limit = max(1, min(200, int(limit)))
    except Exception:
        limit = 50

    out = []
    for f in qs[:limit]:
        meta = (f.payload or {}).get("meta") or {}
        totals = (f.payload or {}).get("totals") or {}
        out.append({
            "id": f.id,
            "oid": f.oid,
            "dt_from": meta.get("dt_from") or "",
            "dt_to": meta.get("dt_to") or "",
            "created_at": f.created_at.isoformat(),
            "toаtals": totals,
        })

    return JsonResponse({"forms": out})


def forms_get(request, form_id: int):
    try:
        f = PutevoyForm.objects.get(id=form_id)
    except PutevoyForm.DoesNotExist:
        raise Http404("Form not found")
    return JsonResponse({"id": f.id, "payload": f.payload})


def forms_export_xlsx(request, form_id: int):
    try:
        f = PutevoyForm.objects.get(id=form_id)
    except PutevoyForm.DoesNotExist:
        raise Http404("Form not found")

    template_path = "/opt/volovo_django/Камаз-маз.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    payload = f.payload or {}
    meta = payload.get("meta") or {}
    totals = payload.get("totals") or {}
    rows = payload.get("rows") or []

    def safe_set(addr: str, value):
        """
        Пишет в ячейку, но если она внутри merged-range — пишет в верхнюю левую.
        Иначе openpyxl даст: MergedCell value is read-only
        """
        cell = ws[addr]
        for r in ws.merged_cells.ranges:
            if cell.coordinate in r:
                ws.cell(row=r.min_row, column=r.min_col).value = value
                return
        cell.value = value

    # --- META: пишем в "Особые отметки" (B66) — безопасное место
    safe_set(
        "B66",
        f"OID={meta.get('oid','')}; "
        f"С={meta.get('dt_from','')}; "
        f"По={meta.get('dt_to','')}"
    )

    # --- Таблица "ПОСЛЕДОВАТЕЛЬНОСТЬ..." начинается с 56 строки (8 строк до 63)
    # Колонки по шаблону (из твоего dump):
    # B  — пункт погрузки/разгрузки (route)
    # Q  — № ездки (tripNo)
    # W  — км
    # AG — тонн
    # AZ — "36" (у тебя delivery)
    # BJ — "38" (у тебя idle)
    start_row = 56

    for i in range(8):
        rr = start_row + i
        r = rows[i] if i < len(rows) else {}

        # очистим ключевые поля (чтобы не оставалось "8" и т.п.)
        safe_set(f"B{rr}", "")
        safe_set(f"Q{rr}", "")
        safe_set(f"W{rr}", "")
        safe_set(f"AG{rr}", "")
        safe_set(f"AZ{rr}", "")
        safe_set(f"BJ{rr}", "")

        # заполняем
        safe_set(f"B{rr}",  r.get("route", "") or "")
        safe_set(f"Q{rr}",  str(i + 1))  # ✅ всегда номер ездки 1..8
        safe_set(f"W{rr}",  r.get("km", "") or "")
        safe_set(f"AG{rr}", r.get("tons", "") or "")
        safe_set(f"AZ{rr}", r.get("delivery", "") or "")
        safe_set(f"BJ{rr}", r.get("idle", "") or "")
    # --- Итоги пока пишем тоже в "Особые отметки" (чтобы не гадать по ячейкам)
    totals_text = (
        f"ИТОГО: km_spread={totals.get('km_spread','')}; "
        f"tons_sum={totals.get('tons_sum','')}; "
        f"km_gps={totals.get('km_gps','')}; "
        f"delivery={totals.get('delivery','')}; "
        f"idle={totals.get('idle','')}"
    )
    # допишем в соседнюю строку, чтобы не мешалось
    safe_set("B67", totals_text)

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        open(tmp.name, "rb"),
        as_attachment=True,
        filename=f"putevoy-{form_id}.xlsx",
    )
