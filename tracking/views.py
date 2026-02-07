import json
from django.http import JsonResponse, HttpResponseBadRequest, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook

from formsapp.models import PutevoyForm


def _parse_dt_from_input(v: str):
    # из html datetime-local приходит "YYYY-MM-DDTHH:MM"
    if not v:
        return None
    s = str(v).strip().replace("T", " ")
    if len(s) == 16:
        s += ":00"
    return parse_datetime(s)


@csrf_exempt
def forms_save(request):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    meta = payload.get("meta") or {}
    oid = meta.get("oid") or None

    form = PutevoyForm.objects.create(
        oid=int(oid) if str(oid).isdigit() else None,
        dt_from=_parse_dt_from_input(meta.get("dt_from", "")),
        dt_to=_parse_dt_from_input(meta.get("dt_to", "")),
        payload=payload,
    )
    return JsonResponse({"form_id": form.id})


def forms_list(request):
    qs = PutevoyForm.objects.all()

    oid = request.GET.get("oid")
    if oid and oid.isdigit():
        qs = qs.filter(oid=int(oid))

    limit = request.GET.get("limit", "50")
    try:
        limit = max(1, min(200, int(limit)))
    except Exception:
        limit = 50

    out = []
    for f in qs[:limit]:
        meta = (f.payload or {}).get("meta") or {}
        totals = (f.payload or {}).get("totals") or {}
        out.append({
            "id": f.id,
            "oid": f.oid,
            "dt_from": meta.get("dt_from") or "",
            "dt_to": meta.get("dt_to") or "",
            "created_at": f.created_at.isoformat(),
            "totals": totals,
        })

    return JsonResponse({"forms": out})


def forms_get(request, form_id: int):
    try:
        f = PutevoyForm.objects.get(id=form_id)
    except PutevoyForm.DoesNotExist:
        raise Http404("Form not found")
    return JsonResponse({"id": f.id, "payload": f.payload})


def forms_export_xlsx(request, form_id: int):
    try:
        f = PutevoyForm.objects.get(id=form_id)
    except PutevoyForm.DoesNotExist:
        raise Http404("Form not found")

    template_path = "/opt/volovo_django/Камаз-маз.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    payload = f.payload or {}
    meta = payload.get("meta") or {}
    totals = payload.get("totals") or {}
    rows = payload.get("rows") or []

    # ⚠️ ВАЖНО: координаты ячеек в шаблоне могут отличаться.
    # Это рабочая "заготовка" — потом точечно попадём в нужные клетки.
    ws["B2"] = f"OID: {meta.get('oid','')}"
    ws["B3"] = f"С: {meta.get('dt_from','')}"
    ws["B4"] = f"По: {meta.get('dt_to','')}"

    start_row = 10
    for i, r in enumerate(rows[:8]):
        rr = start_row + i
        ws[f"A{rr}"] = r.get("route", "")
        ws[f"B{rr}"] = r.get("tripNo", "")
        ws[f"C{rr}"] = r.get("km", "")
        ws[f"D{rr}"] = r.get("tons", "")
        ws[f"E{rr}"] = r.get("width", "")
        ws[f"F{rr}"] = r.get("length", "")
        ws[f"G{rr}"] = r.get("pssTonnage", "")
        ws[f"H{rr}"] = r.get("delivery", "")
        # если нужно
        if "idle" in r:
            ws[f"I{rr}"] = r.get("idle", "")

    ws["C20"] = totals.get("km_spread", "")
    ws["D20"] = totals.get("tons_sum", "")
    ws["E20"] = totals.get("km_gps", "")
    ws["F20"] = totals.get("delivery", "")
    ws["G20"] = totals.get("idle", "")

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(open(tmp.name, "rb"), as_attachment=True, filename=f"putevoy-{form_id}.xlsx")
